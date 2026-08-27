from fastapi import APIRouter, Depends, Query, Response, Body
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import date, timedelta, datetime
import csv
import openpyxl

from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import base64
import os
import tempfile
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import io

from api.deps import get_db
import models
from schemas.reports import (
    FinancialBreakdownItem, FinancialReportSummary, FinancialTrendPoint, FinancialReportResponse,
    SalesReportResponse, SalesReportSummary, SalesTransaction, SalesTrendPoint, PaymentMethodStats, TopMedicineStats,
    PurchaseReportResponse, PurchaseReportSummary, PurchaseTransaction, PurchaseTrendPoint, SupplierStats, TopPurchasedMedicineStats,
    InventoryReportResponse, InventoryReportSummary, InventoryMovementSummary, InventoryStockItem, StockValueByCategory, MedicineMovementItem,
    MedicineReportResponse, MedicineReportSummary, MedicineExpiryItem, MedicineLowStockItem, MedicineMovementAnalyticsItem, PaginationMetadata
)


from pydantic import BaseModel
from typing import Optional

class PDFExportRequest(BaseModel):
    timeframe: str = 'this_month'
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    chart_image: Optional[str] = None
    customer_id: Optional[str] = None
    payment_method: Optional[str] = None
    supplier_id: Optional[str] = None
    report_type: Optional[str] = 'expiry'

router = APIRouter()

def get_reports_date_range(timeframe: str, start_date: str = None, end_date: str = None):
    today = date.today()
    if timeframe == 'custom' and start_date and end_date:
        try:
            sd = datetime.strptime(start_date, '%Y-%m-%d').date()
            ed = datetime.strptime(end_date, '%Y-%m-%d').date()
            return sd, ed
        except ValueError:
            return today, today
    elif timeframe == 'today':
        return today, today
    elif timeframe == 'yesterday':
        yesterday = today - timedelta(days=1)
        return yesterday, yesterday
    elif timeframe == 'last_7_days' or timeframe == 'week':
        return today - timedelta(days=6), today
    elif timeframe == 'last_30_days' or timeframe == 'month':
        return today - timedelta(days=29), today
    elif timeframe == 'this_month':
        return today.replace(day=1), today
    elif timeframe == 'last_month':
        first_day_this_month = today.replace(day=1)
        last_day_last_month = first_day_this_month - timedelta(days=1)
        first_day_last_month = last_day_last_month.replace(day=1)
        return first_day_last_month, last_day_last_month
    elif timeframe == 'this_year' or timeframe == 'year':
        return today.replace(month=1, day=1), today
    else:
        return today, today

def generate_trend_sequence(start_date: date, end_date: date, interval: str):
    dates = []
    if interval == 'hourly':
        for i in range(24):
            dates.append(f"{i:02d}:00")
    elif interval == 'monthly':
        current_date = start_date.replace(day=1)
        while current_date <= end_date:
            dates.append(current_date.strftime('%Y-%m'))
            next_month = current_date.month % 12 + 1
            next_year = current_date.year + (current_date.month // 12)
            current_date = current_date.replace(year=next_year, month=next_month, day=1)
    else:
        current_date = start_date
        while current_date <= end_date:
            dates.append(current_date.strftime('%Y-%m-%d'))
            current_date += timedelta(days=1)
    return dates

def fetch_sales_report_data(
    db: Session, 
    start_date: date, 
    end_date: date, 
    customer_id: str = None, 
    payment_method: str = None
):
    base_query = db.query(models.Sale).filter(
        func.date(models.Sale.TransactionDate) >= start_date,
        func.date(models.Sale.TransactionDate) <= end_date
    )
    if customer_id and customer_id != "all":
        base_query = base_query.filter(models.Sale.CustomerId == customer_id)
    if payment_method and payment_method != "all":
        base_query = base_query.filter(models.Sale.PaymentMethod == payment_method)

    completed_sales = base_query.filter(models.Sale.Status == 'Completed').all()
    # Assuming Sales Returns/Refunds are marked as 'Returned' or tracked elsewhere. 
    # For now, let's say 'Returned' status means full return.
    returned_sales = base_query.filter(models.Sale.Status == 'Returned').all()

    total_gross_sales = sum(float(s.GrandTotal or 0.0) for s in completed_sales)
    total_returns = sum(float(s.GrandTotal or 0.0) for s in returned_sales)
    net_sales = total_gross_sales - total_returns

    # Calculate COGS dynamically
    total_cogs = 0
    for sale in completed_sales:
        for item in sale.items:
            # item.BatchId -> get batch
            batch = db.query(models.StockBatch).filter(models.StockBatch.BatchId == item.BatchId).first()
            if batch:
                total_cogs += (item.Quantity * float(batch.CostPrice or 0.0))

    net_profit = net_sales - total_cogs
    profit_margin = (net_profit / net_sales * 100) if net_sales > 0 else 0.0
    total_invoices = len(completed_sales)
    average_sale = (net_sales / total_invoices) if total_invoices > 0 else 0.0
    highest_sale = max([float(s.GrandTotal or 0.0) for s in completed_sales], default=0.0)

    summary = SalesReportSummary(
        TotalGrossSales=total_gross_sales,
        TotalReturns=total_returns,
        NetSales=net_sales,
        TotalCOGS=total_cogs,
        NetProfit=net_profit,
        ProfitMarginPercent=profit_margin,
        TotalInvoices=total_invoices,
        AverageSale=average_sale,
        HighestSale=highest_sale
    )

    transactions = []
    for s in completed_sales:
        customer_name = s.customer.FirstName + " " + s.customer.LastName if s.customer else "Walk-in"
        total_qty = sum(i.Quantity for i in s.items)
        transactions.append(SalesTransaction(
            InvoiceNo=s.InvoiceNumber or str(s.SalesId),
            TransactionDate=s.TransactionDate,
            CustomerName=customer_name,
            MedicinesSold=len(s.items),
            TotalQty=total_qty,
            Discount=s.DiscountAmount,
            Tax=s.TaxAmount,
            GrandTotal=float(s.GrandTotal or 0.0),
            PaymentMethod=s.PaymentMethod,
            Status=s.Status
        ))

    # Trend Data
    days_diff = (end_date - start_date).days
    if days_diff == 0:
        interval = 'hourly'
        fmt = '%H:00'
    elif days_diff > 60:
        interval = 'monthly'
        fmt = '%Y-%m'
    else:
        interval = 'daily'
        fmt = '%Y-%m-%d'

    date_seq = generate_trend_sequence(start_date, end_date, interval)
    trend_dict = {d: {"sales": 0.0, "cogs": 0.0} for d in date_seq}

    for s in completed_sales:
        period = s.TransactionDate.strftime(fmt)
        if period in trend_dict:
            trend_dict[period]["sales"] += float(s.GrandTotal or 0.0)
            sale_cogs = sum((i.Quantity * float(db.query(models.StockBatch).filter(models.StockBatch.BatchId == i.BatchId).first().CostPrice if db.query(models.StockBatch).filter(models.StockBatch.BatchId == i.BatchId).first() else 0.0)) for i in s.items)
            trend_dict[period]["cogs"] += sale_cogs

    trend_data = []
    for d in date_seq:
        s = trend_dict[d]["sales"]
        c = trend_dict[d]["cogs"]
        p = s - c
        trend_data.append(SalesTrendPoint(label=d, sales=s, profit=p))

    # Payment Methods
    pm_dict = {}
    for s in completed_sales:
        pm = s.PaymentMethod or 'Unknown'
        pm_dict[pm] = pm_dict.get(pm, 0) + float(s.GrandTotal or 0.0)
    payment_methods = [PaymentMethodStats(name=k, value=v) for k, v in pm_dict.items()]

    # Top Medicines
    med_dict = {}
    for s in completed_sales:
        for i in s.items:
            batch = db.query(models.StockBatch).filter(models.StockBatch.BatchId == i.BatchId).first()
            if batch and batch.medicine:
                name = batch.medicine.BrandName
                if name not in med_dict:
                    med_dict[name] = {"qty": 0, "rev": 0.0}
                med_dict[name]["qty"] += i.Quantity
                med_dict[name]["rev"] += float(i.TotalPrice or 0.0)

    top_meds_sorted = sorted(med_dict.items(), key=lambda x: x[1]["rev"], reverse=True)[:5]
    top_medicines = [TopMedicineStats(name=k, quantity=v["qty"], revenue=v["rev"]) for k, v in top_meds_sorted]

    return SalesReportResponse(
        summary=summary,
        trend_data=trend_data,
        payment_methods=payment_methods,
        top_medicines=top_medicines,
        transactions=transactions
    )

@router.get("/sales")
def get_sales_report(
    timeframe: str = 'this_month',
    start_date: str = None,
    end_date: str = None,
    customer_id: str = None,
    payment_method: str = None,
    db: Session = Depends(get_db)
):
    try:
        sd, ed = get_reports_date_range(timeframe, start_date, end_date)
        return fetch_sales_report_data(db, sd, ed, customer_id, payment_method)
    except Exception as e:
        from fastapi.responses import JSONResponse
        return JSONResponse(status_code=500, content={"success": False, "error": str(e)})

@router.get("/sales/export/csv")
def export_sales_report_csv(
    timeframe: str = 'this_month',
    start_date: str = None,
    end_date: str = None,
    customer_id: str = None,
    payment_method: str = None,
    db: Session = Depends(get_db)
):
    sd, ed = get_reports_date_range(timeframe, start_date, end_date)
    report_data = fetch_sales_report_data(db, sd, ed, customer_id, payment_method)
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Invoice No', 'Date', 'Customer', 'Medicines Sold', 'Total Qty', 'Discount', 'Tax', 'Grand Total', 'Payment Method', 'Status'])
    
    for t in report_data.transactions:
        writer.writerow([
            t.InvoiceNo, 
            t.TransactionDate.strftime("%Y-%m-%d %H:%M:%S"),
            t.CustomerName,
            t.MedicinesSold,
            t.TotalQty,
            round(t.Discount, 2),
            round(t.Tax, 2),
            round(t.GrandTotal, 2),
            t.PaymentMethod,
            t.Status
        ])
    
    response = Response(content=output.getvalue(), media_type="text/csv")
    response.headers["Content-Disposition"] = f"attachment; filename=sales_report_{sd}_to_{ed}.csv"
    return response

def fetch_purchase_report_data(
    db: Session, 
    start_date: date, 
    end_date: date, 
    supplier_id: str = None
):
    base_query = db.query(models.Purchase).filter(
        func.date(models.Purchase.PurchaseDate) >= start_date,
        func.date(models.Purchase.PurchaseDate) <= end_date
    )
    if supplier_id and supplier_id != "all":
        base_query = base_query.filter(models.Purchase.SupplierId == supplier_id)

    completed_purchases = base_query.all()
    # Assuming Purchase Returns are tracked via PaymentStatus or elsewhere
    returned_purchases = base_query.filter(models.Purchase.PaymentStatus == 'Returned').all()

    total_gross_purchases = sum(float(p.GrandTotal or 0.0) for p in completed_purchases)
    total_returns = sum(float(p.GrandTotal or 0.0) for p in returned_purchases)
    net_purchases = total_gross_purchases - total_returns

    total_invoices = len(completed_purchases)
    average_purchase = (net_purchases / total_invoices) if total_invoices > 0 else 0.0
    highest_purchase = max([float(p.GrandTotal or 0.0) for p in completed_purchases], default=0.0)

    summary = PurchaseReportSummary(
        TotalGrossPurchases=total_gross_purchases,
        TotalReturns=total_returns,
        NetPurchases=net_purchases,
        TotalInvoices=total_invoices,
        AveragePurchase=average_purchase,
        HighestPurchase=highest_purchase
    )

    transactions = []
    for p in completed_purchases:
        supplier_name = p.supplier.Name if p.supplier else "Unknown"
        total_qty = sum(i.Quantity for i in p.items)
        transactions.append(PurchaseTransaction(
            InvoiceNo=p.InvoiceNumber or str(p.PurchaseId),
            PurchaseDate=p.PurchaseDate,
            SupplierName=supplier_name,
            MedicinesPurchased=len(p.items),
            TotalQty=total_qty,
            Discount=p.TotalDiscount,
            Tax=p.TotalTax,
            GrandTotal=float(p.GrandTotal or 0.0),
            Status=p.PaymentStatus
        ))

    # Trend Data
    days_diff = (end_date - start_date).days
    if days_diff == 0:
        interval = 'hourly'
        fmt = '%H:00'
    elif days_diff > 60:
        interval = 'monthly'
        fmt = '%Y-%m'
    else:
        interval = 'daily'
        fmt = '%Y-%m-%d'

    date_seq = generate_trend_sequence(start_date, end_date, interval)
    trend_dict = {d: 0.0 for d in date_seq}

    for p in completed_purchases:
        period = p.PurchaseDate.strftime(fmt)
        if period in trend_dict:
            trend_dict[period] += float(p.GrandTotal or 0.0)

    trend_data = []
    for d in date_seq:
        trend_data.append(PurchaseTrendPoint(label=d, purchases=trend_dict[d]))

    # Suppliers Breakdown
    supp_dict = {}
    for p in completed_purchases:
        s_name = p.supplier.Name if p.supplier else 'Unknown'
        supp_dict[s_name] = supp_dict.get(s_name, 0) + float(p.GrandTotal or 0.0)
    suppliers = [SupplierStats(name=k, value=v) for k, v in supp_dict.items()]

    # Top Medicines
    med_dict = {}
    for p in completed_purchases:
        for i in p.items:
            med = db.query(models.Medicine).filter(models.Medicine.MedicineId == i.MedicineId).first()
            name = med.BrandName if med else "Unknown"
            if name not in med_dict:
                med_dict[name] = {"qty": 0, "cost": 0.0}
            med_dict[name]["qty"] += i.Quantity
            med_dict[name]["cost"] += float(i.LineTotal or 0.0)

    top_meds_sorted = sorted(med_dict.items(), key=lambda x: x[1]["cost"], reverse=True)[:5]
    top_medicines = [TopPurchasedMedicineStats(name=k, quantity=v["qty"], cost=v["cost"]) for k, v in top_meds_sorted]

    return PurchaseReportResponse(
        summary=summary,
        trend_data=trend_data,
        suppliers=suppliers,
        top_medicines=top_medicines,
        transactions=transactions
    )

@router.get("/purchases")
def get_purchase_report(
    timeframe: str = 'this_month',
    start_date: str = None,
    end_date: str = None,
    supplier_id: str = None,
    db: Session = Depends(get_db)
):
    try:
        sd, ed = get_reports_date_range(timeframe, start_date, end_date)
        return fetch_purchase_report_data(db, sd, ed, supplier_id)
    except Exception as e:
        from fastapi.responses import JSONResponse
        return JSONResponse(status_code=500, content={"success": False, "error": str(e)})

@router.get("/purchases/export/csv")
def export_purchase_report_csv(
    timeframe: str = 'this_month',
    start_date: str = None,
    end_date: str = None,
    supplier_id: str = None,
    db: Session = Depends(get_db)
):
    sd, ed = get_reports_date_range(timeframe, start_date, end_date)
    report_data = fetch_purchase_report_data(db, sd, ed, supplier_id)
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Invoice No', 'Date', 'Supplier', 'Medicines Purchased', 'Total Qty', 'Discount', 'Tax', 'Grand Total', 'Status'])
    
    for t in report_data.transactions:
        writer.writerow([
            t.InvoiceNo, 
            t.PurchaseDate.strftime("%Y-%m-%d %H:%M:%S"),
            t.SupplierName,
            t.MedicinesPurchased,
            t.TotalQty,
            round(t.Discount, 2),
            round(t.Tax, 2),
            round(t.GrandTotal, 2),
            t.Status
        ])
    
    response = Response(content=output.getvalue(), media_type="text/csv")
    response.headers["Content-Disposition"] = f"attachment; filename=purchase_report_{sd}_to_{ed}.csv"
    return response

def fetch_inventory_report_data(
    db: Session,
    start_date: date = None,
    end_date: date = None
):
    today = date.today()
    # 1. Snapshot metrics
    batches = db.query(models.StockBatch).all()
    
    total_cost_value = 0.0
    total_retail_value = 0.0
    expired_valuation = 0.0
    
    # Pre-fetch medicines to avoid N+1 and get Category info
    medicines_db = db.query(models.Medicine).all()
    medicine_map = {m.MedicineId: m for m in medicines_db}
    
    categories = db.query(models.Category).all()
    cat_map = {c.CategoryId: c for c in categories}
    
    low_stock_count = 0
    out_of_stock_count = 0
    
    # Calculate stock per medicine to determine low/out of stock
    med_stock_map = {m.MedicineId: 0 for m in medicines_db}
    
    stock_items = []
    category_valuation_map = {}
    
    for b in batches:
        med = medicine_map.get(b.MedicineId)
        if not med:
            continue
            
        qty = b.Quantity
        cost = float(b.CostPrice)
        sell = float(b.SellingPrice)
        expiry = b.ExpiryDate
        
        is_expired = expiry < today
        val_cost = qty * cost
        val_retail = qty * sell
        
        if is_expired:
            expired_valuation += val_cost
            status = 'Expired'
        else:
            total_cost_value += val_cost
            total_retail_value += val_retail
            status = 'Active'
            med_stock_map[b.MedicineId] += qty
            
            # Category Valuation (only active sellable)
            cat_name = cat_map[med.CategoryId].CategoryName if med.CategoryId in cat_map else 'Unknown'
            category_valuation_map[cat_name] = category_valuation_map.get(cat_name, 0.0) + val_cost
            
        stock_items.append(InventoryStockItem(
            MedicineName=med.BrandName,
            Category=cat_map[med.CategoryId].CategoryName if med.CategoryId in cat_map else 'Unknown',
            BatchCode=b.BatchCode,
            Quantity=qty,
            CostPrice=cost,
            SellingPrice=sell,
            TotalCostValue=val_cost,
            TotalRetailValue=val_retail,
            ExpiryDate=expiry,
            Status=status
        ))
    
    for m in medicines_db:
        qty = med_stock_map[m.MedicineId]
        if qty == 0:
            out_of_stock_count += 1
        elif qty <= m.ReorderLevel:
            low_stock_count += 1
            
    summary = InventoryReportSummary(
        TotalCostValue=total_cost_value,
        TotalRetailValue=total_retail_value,
        ExpiredWrittenOffValuation=expired_valuation,
        TotalItemsInStock=sum(med_stock_map.values()),
        LowStockCount=low_stock_count,
        OutOfStockCount=out_of_stock_count
    )
    
    category_valuation = [StockValueByCategory(name=k, value=v) for k, v in category_valuation_map.items()]
    
    # 2. Movement metrics (if date range provided)
    movement_summary = None
    movement_items = None
    
    if start_date and end_date:
        purchases = db.query(models.PurchaseItem).join(models.Purchase).filter(
            func.date(models.Purchase.PurchaseDate) >= start_date,
            func.date(models.Purchase.PurchaseDate) <= end_date
        ).all()
        
        sales = db.query(models.SaleItem).join(models.Sale).filter(
            func.date(models.Sale.TransactionDate) >= start_date,
            func.date(models.Sale.TransactionDate) <= end_date
        ).all()
        
        adjustments = db.query(models.StockAdjustment).filter(
            func.date(models.StockAdjustment.AdjustmentDate) >= start_date,
            func.date(models.StockAdjustment.AdjustmentDate) <= end_date
        ).all()
        
        # We don't have an explicit 'expired qty' in the movement tables, 
        # but adjustments might have Reason='Expired'.
        # We'll treat adjustments with Reason='Expired' or 'Damage' as Expired/Written-Off
        
        tot_purchased = 0
        tot_sold = 0
        tot_adjusted = 0
        tot_expired_writeoff = 0
        
        med_move_map = {m.MedicineId: {"p": 0, "s": 0, "a": 0, "e": 0} for m in medicines_db}
        
        for p in purchases:
            tot_purchased += p.Quantity
            med_move_map[p.MedicineId]["p"] += p.Quantity
            
        for s in sales:
            tot_sold += s.Quantity
            # s.BatchId -> get medicine
            b = db.query(models.StockBatch).filter(models.StockBatch.BatchId == s.BatchId).first()
            if b:
                med_move_map[b.MedicineId]["s"] += s.Quantity
                
        for a in adjustments:
            b = db.query(models.StockBatch).filter(models.StockBatch.BatchId == a.BatchId).first()
            if not b:
                continue
                
            qty = a.Quantity if a.AdjustmentType == 'Increase' else -a.Quantity
            reason = a.Reason.lower()
            
            if 'expir' in reason or 'damag' in reason or 'write' in reason:
                # Typically negative for expiry write-offs
                tot_expired_writeoff += abs(qty)
                med_move_map[b.MedicineId]["e"] += abs(qty)
            else:
                tot_adjusted += qty
                med_move_map[b.MedicineId]["a"] += qty
                
        movement_summary = InventoryMovementSummary(
            PurchasedQty=tot_purchased,
            SoldQty=tot_sold,
            ManualAdjustmentsQty=tot_adjusted,
            ExpiredWrittenOffQty=tot_expired_writeoff
        )
        
        movement_items = []
        for m in medicines_db:
            m_move = med_move_map[m.MedicineId]
            if m_move["p"] == 0 and m_move["s"] == 0 and m_move["a"] == 0 and m_move["e"] == 0:
                continue # Skip if no movement
                
            # Current stock
            current_stock = med_stock_map[m.MedicineId]
            # Starting stock = Current - Additions + Reductions
            # Additions = Purchased + Positive Adjustments
            # Reductions = Sold + Negative Adjustments + Expired Writeoffs
            # This is a simplification (ignores returns if they aren't included in sales/purchases above).
            
            closing_stock = current_stock
            starting_stock = closing_stock - m_move["p"] + m_move["s"] - m_move["a"] + m_move["e"]
            
            movement_items.append(MedicineMovementItem(
                MedicineName=m.BrandName,
                StartingStock=starting_stock,
                PurchasedQty=m_move["p"],
                SoldQty=m_move["s"],
                AdjustedQty=m_move["a"],
                ExpiredQty=m_move["e"],
                ClosingStock=closing_stock
            ))
    
    return InventoryReportResponse(
        summary=summary,
        movement_summary=movement_summary,
        stock_items=stock_items,
        movement_items=movement_items,
        category_valuation=category_valuation
    )

@router.get("/inventory")
def get_inventory_report(
    timeframe: str = None,
    start_date: str = None,
    end_date: str = None,
    db: Session = Depends(get_db)
):
    try:
        sd, ed = None, None
        if timeframe and timeframe != 'all':
            sd, ed = get_reports_date_range(timeframe, start_date, end_date)
            
        return fetch_inventory_report_data(db, sd, ed)
    except Exception as e:
        from fastapi.responses import JSONResponse
        return JSONResponse(status_code=500, content={"success": False, "error": str(e)})

@router.get("/inventory/export/csv")
def export_inventory_report_csv(
    db: Session = Depends(get_db)
):
    # For export, we typically just export current stock
    report_data = fetch_inventory_report_data(db, None, None)
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Medicine', 'Category', 'Batch Code', 'Quantity', 'Cost Price', 'Selling Price', 'Total Cost Value', 'Total Retail Value', 'Expiry Date', 'Status'])
    
    for t in report_data.stock_items:
        writer.writerow([
            t.MedicineName,
            t.Category,
            t.BatchCode,
            t.Quantity,
            round(t.CostPrice, 2),
            round(t.SellingPrice, 2),
            round(t.TotalCostValue, 2),
            round(t.TotalRetailValue, 2),
            t.ExpiryDate.strftime("%Y-%m-%d") if t.ExpiryDate else '',
            t.Status
        ])
    
    response = Response(content=output.getvalue(), media_type="text/csv")
    response.headers["Content-Disposition"] = f"attachment; filename=inventory_report.csv"
    return response


def fetch_medicine_report_data(
    db: Session, 
    start_date: date, 
    end_date: date,
    report_type: str = None,
    search: str = None,
    category_id: int = None,
    page: int = 1,
    page_size: int = 10
):
    from sqlalchemy import select
    today = date.today()
    
    # --- 1. Expiry Items (Batch-level) ---
    expiry_items = []
    total_expired = 0
    expiring_soon = 0
    
    q_expiry = db.query(
        models.StockBatch, 
        models.Medicine, 
        models.Supplier.Name.label('supplier_name')
    ).join(models.Medicine, models.StockBatch.MedicineId == models.Medicine.MedicineId)\
     .outerjoin(
        models.PurchaseItem, 
        (models.PurchaseItem.MedicineId == models.StockBatch.MedicineId) & 
        (models.PurchaseItem.BatchCode == models.StockBatch.BatchCode)
     ).outerjoin(models.Purchase, models.PurchaseItem.PurchaseId == models.Purchase.PurchaseId)\
     .outerjoin(models.Supplier, models.Purchase.SupplierId == models.Supplier.SupplierId)\
     .filter(models.StockBatch.Quantity > 0)
     
    if category_id:
        q_expiry = q_expiry.filter(models.Medicine.CategoryId == category_id)
    if search:
        q_expiry = q_expiry.filter(
            (models.Medicine.BrandName.ilike(f"%{search}%")) | 
            (models.StockBatch.BatchCode.ilike(f"%{search}%"))
        )
        
    seen_batches = set()
    for batch, medicine, supplier_name in q_expiry.all():
        if batch.BatchId in seen_batches:
            continue
        seen_batches.add(batch.BatchId)
        
        if not batch.ExpiryDate: continue
        days_to_expiry = (batch.ExpiryDate - today).days
        status = ''
        if days_to_expiry < 0:
            status = 'Expired'
            total_expired += 1
        elif days_to_expiry <= 30:
            status = 'Expiring < 30 days'
            expiring_soon += 1
        elif days_to_expiry <= 90:
            status = 'Expiring < 90 days'
            expiring_soon += 1
        else:
            status = 'Safe'
            
        if report_type == 'expiry' and start_date and end_date:
            if batch.ExpiryDate < start_date or batch.ExpiryDate > end_date:
                continue

        if status != 'Safe':
            expiry_items.append(MedicineExpiryItem(
                MedicineName=medicine.BrandName,
                BatchCode=batch.BatchCode,
                Quantity=batch.Quantity,
                ExpiryDate=batch.ExpiryDate,
                DaysToExpiry=days_to_expiry,
                Status=status,
                SupplierName=supplier_name
            ))
            
    expiry_items.sort(key=lambda x: x.DaysToExpiry)

    # --- 2. Medicine-level (Low Stock & Performance) ---
    subq = select(models.Supplier.Name).select_from(models.PurchaseItem)\
        .join(models.Purchase)\
        .join(models.Supplier)\
        .where(models.PurchaseItem.MedicineId == models.Medicine.MedicineId)\
        .order_by(models.Purchase.PurchaseDate.desc())\
        .limit(1).scalar_subquery()

    q_meds = db.query(
        models.Medicine,
        models.Category,
        subq.label('supplier_name')
    ).outerjoin(models.Category, models.Medicine.CategoryId == models.Category.CategoryId)

    if category_id:
        q_meds = q_meds.filter(models.Medicine.CategoryId == category_id)
    if search:
        q_meds = q_meds.filter(models.Medicine.BrandName.ilike(f"%{search}%"))
        
    medicines_data = q_meds.all()
    
    # 2a. Low Stock
    low_stock_items = []
    active_medicines = set()
    for med, cat, supplier_name in medicines_data:
        current_stock = sum(b.Quantity for b in med.batches if b.Quantity > 0)
        if current_stock > 0:
            active_medicines.add(med.MedicineId)
            
        reorder_level = med.ReorderLevel or 0
        if current_stock <= reorder_level:
            deficit = reorder_level - current_stock
            suggested = max((reorder_level * 2) - current_stock, 0)
            low_stock_items.append(MedicineLowStockItem(
                MedicineName=med.BrandName,
                Category=cat.CategoryName if cat else 'Uncategorized',
                CurrentStock=current_stock,
                ReorderLevel=reorder_level,
                Deficit=deficit,
                SuggestedReorderQty=suggested,
                SupplierName=supplier_name
            ))
            
    low_stock_items.sort(key=lambda x: x.Deficit, reverse=True)

    # 2b. Moving Items
    days_in_range = (end_date - start_date).days if (start_date and end_date) else 30
    if days_in_range <= 0: days_in_range = 1
    
    movement_items = []
    fast_count = 0
    slow_count = 0
    dead_count = 0
    
    q_sales = db.query(
        models.StockBatch.MedicineId.label('medicine_id'),
        func.sum(models.SaleItem.Quantity).label('sold_qty'),
        func.sum(models.SaleItem.TotalPrice).label('revenue')
    ).select_from(models.SaleItem).join(models.Sale).join(models.StockBatch)
    
    if start_date and end_date:
        q_sales = q_sales.filter(
            func.date(models.Sale.TransactionDate) >= start_date,
            func.date(models.Sale.TransactionDate) <= end_date
        )
    
    sales_data = q_sales.group_by(models.StockBatch.MedicineId).all()
    sales_map = {item.medicine_id: {'sold_qty': item.sold_qty, 'revenue': item.revenue} for item in sales_data}
    
    for med, cat, supplier_name in medicines_data:
        stats = sales_map.get(med.MedicineId, {'sold_qty': 0, 'revenue': 0})
        sold_qty = stats['sold_qty'] or 0
        revenue = stats['revenue'] or 0
        
        velocity = float(sold_qty) / days_in_range
        
        classification = 'Normal'
        if sold_qty == 0 and med.MedicineId in active_medicines and days_in_range >= 60:
            classification = 'Dead Stock'
            dead_count += 1
        elif velocity >= 2.0:
            classification = 'Fast Moving'
            fast_count += 1
        elif velocity < 0.5 and med.MedicineId in active_medicines:
            classification = 'Slow Moving'
            slow_count += 1
            
        if classification != 'Normal':
            movement_items.append(MedicineMovementAnalyticsItem(
                MedicineName=med.BrandName,
                Category=cat.CategoryName if cat else 'Uncategorized',
                SoldQuantity=sold_qty,
                SalesVelocity=round(velocity, 2),
                Revenue=float(revenue),
                Classification=classification,
                SupplierName=supplier_name
            ))
            
    movement_items.sort(key=lambda x: x.SalesVelocity, reverse=True)

    summary = MedicineReportSummary(
        TotalExpiredBatches=total_expired,
        ExpiringSoonBatches=expiring_soon,
        LowStockMedicines=len(low_stock_items),
        FastMovingCount=fast_count,
        SlowMovingCount=slow_count,
        DeadStockCount=dead_count
    )

    # --- Pagination ---
    target_list = []
    if report_type == 'expiry':
        target_list = expiry_items
    elif report_type == 'low_stock':
        target_list = low_stock_items
    elif report_type == 'moving':
        target_list = movement_items
    else:
        # Default behavior: return all without pagination
        return MedicineReportResponse(
            summary=summary,
            expiry_items=expiry_items,
            low_stock_items=low_stock_items,
            movement_items=movement_items,
            pagination=None
        )

    total = len(target_list)
    if page_size > 0:
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        paginated_list = target_list[start_idx:end_idx]
    else:
        paginated_list = target_list

    pagination_meta = PaginationMetadata(total=total, page=page, page_size=page_size)

    return MedicineReportResponse(
        summary=summary,
        expiry_items=paginated_list if report_type == 'expiry' else [],
        low_stock_items=paginated_list if report_type == 'low_stock' else [],
        movement_items=paginated_list if report_type == 'moving' else [],
        pagination=pagination_meta
    )

@router.get("/medicine", response_model=MedicineReportResponse)
def get_medicine_reports(
    report_type: str = Query(None, description="Type of report: expiry, low_stock, or moving"),
    search: str = Query(None),
    category_id: int = Query(None),
    page: int = Query(1),
    page_size: int = Query(10),
    timeframe: str = Query("last_30_days", description="Timeframe for moving items analysis"),
    start_date: str = Query(None),
    end_date: str = Query(None),
    db: Session = Depends(get_db)
):
    sd, ed = get_reports_date_range(timeframe, start_date, end_date)
    return fetch_medicine_report_data(db, sd, ed, report_type, search, category_id, page, page_size)

@router.get("/medicine/export/csv")
def export_medicine_report_csv(
    report_type: str = Query(..., description="Type of report: expiry, low_stock, or moving"),
    search: str = Query(None),
    category_id: int = Query(None),
    timeframe: str = Query("last_30_days"),
    start_date: str = Query(None),
    end_date: str = Query(None),
    db: Session = Depends(get_db)
):
    sd, ed = get_reports_date_range(timeframe, start_date, end_date)
    # Use page_size=0 to fetch all matches without pagination
    data = fetch_medicine_report_data(db, sd, ed, report_type, search, category_id, 1, 0)
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    if report_type == "expiry":
        writer.writerow(["Medicine Name", "Batch Code", "Quantity", "Expiry Date", "Days to Expiry", "Supplier", "Status"])
        for item in data.expiry_items:
            writer.writerow([item.MedicineName, item.BatchCode, item.Quantity, item.ExpiryDate.strftime("%Y-%m-%d"), item.DaysToExpiry, item.SupplierName or "", item.Status])
    elif report_type == "low_stock":
        writer.writerow(["Medicine Name", "Category", "Supplier", "Current Stock", "Reorder Level", "Deficit", "Suggested Reorder Qty"])
        for item in data.low_stock_items:
            writer.writerow([item.MedicineName, item.Category, item.SupplierName or "", item.CurrentStock, item.ReorderLevel, item.Deficit, item.SuggestedReorderQty])
    elif report_type == "moving":
        writer.writerow(["Medicine Name", "Category", "Supplier", "Sold Quantity", "Sales Velocity (units/day)", "Revenue", "Classification"])
        for item in data.movement_items:
            writer.writerow([item.MedicineName, item.Category, item.SupplierName or "", item.SoldQuantity, item.SalesVelocity, item.Revenue, item.Classification])
    else:
        return Response(status_code=400, content="Invalid report type")
        
    response = Response(content=output.getvalue(), media_type="text/csv")
    response.headers["Content-Disposition"] = f"attachment; filename=medicine_{report_type}_report.csv"
    return response

def fetch_financial_report_data(
    db: Session, 
    start_date: date, 
    end_date: date
):
    # 1. Sales Data (Revenue & Discounts & Returns & COGS)
    completed_sales = db.query(models.Sale).filter(
        func.date(models.Sale.TransactionDate) >= start_date,
        func.date(models.Sale.TransactionDate) <= end_date,
        models.Sale.Status == 'Completed'
    ).all()
    
    returned_sales = db.query(models.SaleReturn).filter(
        func.date(models.SaleReturn.ReturnDate) >= start_date,
        func.date(models.SaleReturn.ReturnDate) <= end_date
    ).all()
    
    gross_sales = sum(float(s.SubTotal or 0.0) for s in completed_sales)
    discounts_applied = sum(float(s.DiscountAmount or 0.0) for s in completed_sales)
    sales_returns = sum(float(r.TotalRefundAmount or 0.0) for r in returned_sales)
    
    total_revenue = gross_sales - discounts_applied - sales_returns
    
    total_cogs = 0.0
    for sale in completed_sales:
        for item in sale.items:
            batch = db.query(models.StockBatch).filter(models.StockBatch.BatchId == item.BatchId).first()
            if batch:
                total_cogs += (item.Quantity * float(batch.CostPrice or 0.0))
                
    # 2. Inventory Loss / Expiry Write-Off
    today = date.today()
    expired_batches = db.query(models.StockBatch).filter(
        models.StockBatch.Quantity > 0,
        models.StockBatch.ExpiryDate < today
    ).all()
    
    inventory_loss = sum((b.Quantity * float(b.CostPrice or 0.0)) for b in expired_batches)
    
    # 3. Profits
    gross_profit = total_revenue - total_cogs
    total_expenses = 0.0  # Operating Expenses set to 0.0 for this phase as proposed
    net_profit = gross_profit - total_expenses - inventory_loss
    
    profit_margin = (net_profit / total_revenue * 100) if total_revenue > 0 else 0.0
    
    summary = FinancialReportSummary(
        GrossSales=gross_sales,
        DiscountsApplied=discounts_applied,
        SalesReturns=sales_returns,
        TotalRevenue=total_revenue,
        TotalCOGS=total_cogs,
        InventoryLoss=inventory_loss,
        GrossProfit=gross_profit,
        TotalExpenses=total_expenses,
        NetProfit=net_profit,
        ProfitMargin=profit_margin
    )
    
    income_breakdown = [
        FinancialBreakdownItem(Category="Sales Revenue (Gross)", Amount=gross_sales)
    ]
    
    expense_breakdown = [
        FinancialBreakdownItem(Category="Cost of Goods Sold", Amount=total_cogs),
        FinancialBreakdownItem(Category="Discounts Applied", Amount=discounts_applied),
        FinancialBreakdownItem(Category="Sales Returns / Refunds", Amount=sales_returns),
        FinancialBreakdownItem(Category="Inventory Loss / Expiry Write-Off", Amount=inventory_loss),
        FinancialBreakdownItem(Category="Operating Expenses", Amount=total_expenses)
    ]
    
    # 4. Trend Data
    days_diff = (end_date - start_date).days
    if days_diff == 0:
        interval = 'hourly'
        fmt = '%H:00'
    elif days_diff > 60:
        interval = 'monthly'
        fmt = '%Y-%m'
    else:
        interval = 'daily'
        fmt = '%Y-%m-%d'

    date_seq = generate_trend_sequence(start_date, end_date, interval)
    trend_dict = {d: {"revenue": 0.0, "expenses": 0.0} for d in date_seq}

    # Add daily revenue
    for s in completed_sales:
        period = s.TransactionDate.strftime(fmt)
        if period in trend_dict:
            # Net revenue for the sale (SubTotal - Discount)
            trend_dict[period]["revenue"] += (float(s.SubTotal or 0.0) - float(s.DiscountAmount or 0.0))
            
            # Add COGS to expenses for the sale
            sale_cogs = sum((i.Quantity * float(db.query(models.StockBatch).filter(models.StockBatch.BatchId == i.BatchId).first().CostPrice if db.query(models.StockBatch).filter(models.StockBatch.BatchId == i.BatchId).first() else 0.0)) for i in s.items)
            trend_dict[period]["expenses"] += sale_cogs
            
    # Add daily returns to expenses (as a reduction of revenue)
    for r in returned_sales:
        period = r.ReturnDate.strftime(fmt)
        if period in trend_dict:
            trend_dict[period]["expenses"] += float(r.TotalRefundAmount or 0.0)

    # Note: Inventory loss is static (current active expired stock), so we distribute it evenly or just skip it in daily trend? 
    # Skipping it in daily trend because it's a cumulative current loss, not realized on a specific day in this date range.
            
    trend_data = []
    for d in date_seq:
        rev = trend_dict[d]["revenue"]
        exp = trend_dict[d]["expenses"]
        prof = rev - exp
        trend_data.append(FinancialTrendPoint(label=d, revenue=rev, expenses=exp, profit=prof))
        
    return FinancialReportResponse(
        summary=summary,
        income_breakdown=income_breakdown,
        expense_breakdown=expense_breakdown,
        trend_data=trend_data
    )

@router.get("/financial", response_model=FinancialReportResponse)
def get_financial_reports(
    timeframe: str = Query("last_30_days"),
    start_date: str = Query(None),
    end_date: str = Query(None),
    db: Session = Depends(get_db)
):
    sd, ed = get_reports_date_range(timeframe, start_date, end_date)
    return fetch_financial_report_data(db, sd, ed)


@router.get("/sales/export/excel")
def export_sales_report_excel(
    timeframe: str = 'this_month',
    start_date: str = None,
    end_date: str = None,
    customer_id: str = None,
    payment_method: str = None,
    db: Session = Depends(get_db)
):
    sd, ed = get_reports_date_range(timeframe, start_date, end_date)
    report_data = fetch_sales_report_data(db, sd, ed, customer_id, payment_method)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    headers = ['Invoice No', 'Date', 'Customer', 'Medicines Sold', 'Total Qty', 'Discount', 'Tax', 'Grand Total', 'Payment Method', 'Status']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    for t in report_data.transactions:
        ws.append([
            t.InvoiceNo, 
            t.TransactionDate.strftime("%Y-%m-%d %H:%M:%S"),
            t.CustomerName,
            t.MedicinesSold,
            t.TotalQty,
            round(t.Discount, 2),
            round(t.Tax, 2),
            round(t.GrandTotal, 2),
            t.PaymentMethod,
            t.Status
        ])
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        
    output = io.BytesIO()
    wb.save(output)
    response = Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename=sales_report_{sd}_to_{ed}.xlsx"
    return response

@router.get("/purchases/export/excel")
def export_purchase_report_excel(
    timeframe: str = 'this_month',
    start_date: str = None,
    end_date: str = None,
    supplier_id: str = None,
    db: Session = Depends(get_db)
):
    sd, ed = get_reports_date_range(timeframe, start_date, end_date)
    report_data = fetch_purchase_report_data(db, sd, ed, supplier_id)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Report"
    
    headers = ['Invoice No', 'Date', 'Supplier', 'Medicines Purchased', 'Total Qty', 'Discount', 'Tax', 'Grand Total', 'Status']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    for t in report_data.transactions:
        ws.append([
            t.InvoiceNo, 
            t.PurchaseDate.strftime("%Y-%m-%d %H:%M:%S"),
            t.SupplierName,
            t.MedicinesPurchased,
            t.TotalQty,
            round(t.Discount, 2),
            round(t.Tax, 2),
            round(t.GrandTotal, 2),
            t.Status
        ])
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = (max_length + 2)
        
    output = io.BytesIO()
    wb.save(output)
    response = Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename=purchase_report_{sd}_to_{ed}.xlsx"
    return response

@router.get("/inventory/export/excel")
def export_inventory_report_excel(
    timeframe: str = 'this_month',
    start_date: str = None,
    end_date: str = None,
    db: Session = Depends(get_db)
):
    sd, ed = get_reports_date_range(timeframe, start_date, end_date)
    report_data = fetch_inventory_report_data(db, sd, ed)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Report"
    
    headers = ['Medicine Name', 'Category', 'Batch Number', 'Stock Quantity', 'Cost Price', 'Selling Price', 'Total Cost Value', 'Total Retail Value', 'Expiry Date', 'Status']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    for t in report_data.stock_items:
        ws.append([
            t.MedicineName,
            t.Category,
            t.BatchCode,
            t.Quantity,
            round(t.CostPrice, 2),
            round(t.SellingPrice, 2),
            round(t.TotalCostValue, 2),
            round(t.TotalRetailValue, 2),
            t.ExpiryDate.strftime("%Y-%m-%d") if getattr(t, 'ExpiryDate', None) else "N/A",
            t.Status
        ])
        
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = (max_length + 2)
        
    output = io.BytesIO()
    wb.save(output)
    response = Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename=inventory_report.xlsx"
    return response

@router.get("/medicine/export/excel")
def export_medicine_report_excel(
    timeframe: str = 'this_month',
    start_date: str = None,
    end_date: str = None,
    report_type: str = 'expiry',
    db: Session = Depends(get_db)
):
    sd, ed = get_reports_date_range(timeframe, start_date, end_date)
    report_data = fetch_medicine_report_data(db, sd, ed)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Medicine {report_type.capitalize()} Report"
    
    if report_type == 'expiry':
        headers = ['Medicine ID', 'Brand Name', 'Batch Number', 'Stock Qty', 'Expiry Date', 'Days to Expire', 'Risk Level']
    elif report_type == 'low_stock':
        headers = ['Medicine ID', 'Brand Name', 'Current Stock', 'Min Stock Level', 'Suggested Reorder Qty', 'Status']
    else:
        headers = ['Medicine ID', 'Brand Name', 'Total Qty Sold', 'Total Revenue', 'Avg Daily Sales', 'Current Stock', 'Classification']
        
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    if report_type == 'expiry':
        items = report_data.expiry_items
    elif report_type == 'low_stock':
        items = report_data.low_stock_items
    else:
        items = report_data.movement_items
        
    for t in items:
        if report_type == 'expiry':
            ws.append([t.MedicineId, t.MedicineName, getattr(t, 'BatchNumber', 'N/A'), getattr(t, 'StockQuantity', 0), getattr(t, 'ExpiryDate', '').strftime("%Y-%m-%d") if getattr(t, 'ExpiryDate', None) else 'N/A', getattr(t, 'DaysToExpiry', 0), getattr(t, 'RiskLevel', 'Unknown')])
        elif report_type == 'low_stock':
            ws.append([getattr(t, 'MedicineId', 'N/A'), t.MedicineName, getattr(t, 'CurrentStock', 0), getattr(t, 'ReorderLevel', 0), getattr(t, 'SuggestedReorderQty', 0), 'Low Stock'])
        else:
            ws.append([getattr(t, 'MedicineId', 'N/A'), t.MedicineName, getattr(t, 'SoldQuantity', 0), round(getattr(t, 'Revenue', 0.0), 2), round(getattr(t, 'SalesVelocity', 0.0), 2), 0, getattr(t, 'Classification', 'Unknown')])
            
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = (max_length + 2)
        
    output = io.BytesIO()
    wb.save(output)
    response = Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename=medicine_{report_type}_report.xlsx"
    return response

@router.get("/financial/export/excel")
def export_financial_report_excel(
    timeframe: str = 'this_month',
    start_date: str = None,
    end_date: str = None,
    db: Session = Depends(get_db)
):
    sd, ed = get_reports_date_range(timeframe, start_date, end_date)
    report_data = fetch_financial_report_data(db, sd, ed)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Report"
    
    headers = ['Category', 'Amount', 'Type']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    for item in report_data.income_breakdown:
        ws.append([item.Category, round(item.Amount, 2), "Income"])
    for item in report_data.expense_breakdown:
        ws.append([item.Category, round(item.Amount, 2), "Expense"])
        
    ws.append([])
    ws.append(["Summary", "Amount"])
    ws.append(["Total Revenue", round(report_data.summary.TotalRevenue, 2)])
    ws.append(["Total COGS", round(report_data.summary.TotalCOGS, 2)])
    ws.append(["Gross Profit", round(report_data.summary.GrossProfit, 2)])
    ws.append(["Total Expenses", round(report_data.summary.TotalExpenses, 2)])
    ws.append(["Net Profit", round(report_data.summary.NetProfit, 2)])
    ws.append(["Profit Margin", f"{round(report_data.summary.ProfitMargin, 2)}%"])
    
    for row in ws.iter_rows(min_row=len(report_data.income_breakdown) + len(report_data.expense_breakdown) + 3, max_row=len(report_data.income_breakdown) + len(report_data.expense_breakdown) + 8):
        row[0].font = Font(bold=True)
        
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = (max_length + 2)
        
    output = io.BytesIO()
    wb.save(output)
    response = Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename=financial_report.xlsx"
    return response




def get_premium_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='PremiumTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=28, textColor=colors.HexColor('#1E293B'), alignment=TA_CENTER, spaceAfter=8))
    styles.add(ParagraphStyle(name='PremiumSubtitle', parent=styles['Normal'], fontName='Helvetica', fontSize=12, textColor=colors.HexColor('#64748B'), alignment=TA_CENTER, spaceAfter=20))
    styles.add(ParagraphStyle(name='HeaderLabel', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=11, textColor=colors.HexColor('#475569')))
    styles.add(ParagraphStyle(name='HeaderValue', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=18, textColor=colors.HexColor('#0F172A')))
    return styles

def build_premium_header(title_text, subtitle_text):
    styles = get_premium_styles()
    return [
        Spacer(1, 20),
        Paragraph(title_text, styles['PremiumTitle']),
        Paragraph(subtitle_text, styles['PremiumSubtitle']),
        Spacer(1, 10),
    ]

def build_kpi_table(kpi_data):
    # kpi_data is a list of tuples: (Label, Value, HexColor)
    # We will arrange them in a horizontal grid.
    styles = get_premium_styles()
    
    table_data = [[]]
    for label, value, color in kpi_data:
        cell_data = [
            Paragraph(label, styles['HeaderLabel']),
            Spacer(1, 8),
            Paragraph(str(value), ParagraphStyle(name='Temp', parent=styles['HeaderValue'], textColor=colors.HexColor(color)))
        ]
        table_data[0].append(cell_data)
        
    kpi_table = Table(table_data, colWidths=[500/len(kpi_data)] * len(kpi_data))
    kpi_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
        ('BOX', (0,0), (-1,-1), 1.5, colors.HexColor('#94A3B8')),
        ('INNERGRID', (0,0), (-1,-1), 1, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0,0), (-1,-1), 18),
        ('BOTTOMPADDING', (0,0), (-1,-1), 18),
    ]))
    return [kpi_table, Spacer(1, 30)]

def embed_chart_in_pdf(elements, chart_image_b64):
    if chart_image_b64:
        try:
            if "," in chart_image_b64:
                chart_image_b64 = chart_image_b64.split(",")[1]
            img_data = base64.b64decode(chart_image_b64)
            with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
                tmp.write(img_data)
                tmp_path = tmp.name
            
            img = RLImage(tmp_path)
            
            # Maintain aspect ratio beautifully, don't just stretch
            aspect = img.imageHeight / float(img.imageWidth)
            # Max width is 500, max height is 300
            desired_width = min(500, img.imageWidth)
            desired_height = desired_width * aspect
            
            if desired_height > 250:
                desired_height = 250
                desired_width = desired_height / aspect
                
            img.drawWidth = desired_width
            img.drawHeight = desired_height
            
            elements.append(img)
            elements.append(Spacer(1, 30))
            
        except Exception as e:
            print("Failed to embed chart:", e)


@router.post("/sales/export/pdf")
def export_sales_report_pdf(req: dict = Body(...), db: Session = Depends(get_db)):
    req = PDFExportRequest(**req)
    sd, ed = get_reports_date_range(req.timeframe, req.start_date, req.end_date)
    report_data = fetch_sales_report_data(db, sd, ed, req.customer_id, req.payment_method)
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    elements.extend(build_premium_header("Sales Report", f"Period: {sd} to {ed}"))
    
    kpi_data = [
        ("Gross Sales", f"Rs. {report_data.summary.TotalGrossSales}", "#3B82F6"),
        ("Returns", f"Rs. {report_data.summary.TotalReturns}", "#EF4444"),
        ("Net Sales", f"Rs. {report_data.summary.NetSales}", "#10B981"),
        ("Invoices", str(report_data.summary.TotalInvoices), "#6366F1"),
    ]
    elements.extend(build_kpi_table(kpi_data))
    
    # Chart
    embed_chart_in_pdf(elements, req.chart_image)
    
    # Table Data
    data = [['Invoice No', 'Date', 'Customer', 'Medicines', 'Total Qty', 'Grand Total', 'Status']]
    for t in report_data.transactions:
        data.append([
            t.InvoiceNo, 
            t.TransactionDate.strftime("%Y-%m-%d %H:%M"),
            t.CustomerName[:15], # Truncate long names for PDF fit
            str(t.MedicinesSold),
            str(t.TotalQty),
            str(round(t.GrandTotal, 2)),
            t.Status
        ])
        
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#FFFFFF')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 12),
        ('BOTTOMPADDING', (0,0), (-1,0), 14),
        ('TOPPADDING', (0,0), (-1,0), 14),
        
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#FFFFFF'), colors.HexColor('#F1F5F9')]),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-1), 10),
        ('BOTTOMPADDING', (0,1), (-1,-1), 10),
        ('TOPPADDING', (0,1), (-1,-1), 10),
        
        ('LINEBELOW', (0,0), (-1,0), 2, colors.HexColor('#3B82F6')), 
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('BOX', (0,0), (-1,-1), 1.5, colors.HexColor('#94A3B8')),
    ]))
    elements.append(t)
    
    doc.build(elements)
    response = Response(content=output.getvalue(), media_type="application/pdf")
    response.headers["Content-Disposition"] = f"attachment; filename=sales_report_{sd}_to_{ed}.pdf"
    return response

@router.post("/purchases/export/pdf")
def export_purchase_report_pdf(req: dict = Body(...), db: Session = Depends(get_db)):
    req = PDFExportRequest(**req)
    sd, ed = get_reports_date_range(req.timeframe, req.start_date, req.end_date)
    report_data = fetch_purchase_report_data(db, sd, ed, req.supplier_id)
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    elements.extend(build_premium_header("Purchase Report", f"Period: {sd} to {ed}"))
    
    kpi_data = [
        ("Gross Purchases", f"Rs. {report_data.summary.TotalGrossPurchases}", "#3B82F6"),
        ("Returns", f"Rs. {report_data.summary.TotalReturns}", "#EF4444"),
        ("Net Purchases", f"Rs. {report_data.summary.NetPurchases}", "#10B981"),
        ("Invoices", str(report_data.summary.TotalInvoices), "#6366F1"),
    ]
    elements.extend(build_kpi_table(kpi_data))
    
    # Chart
    embed_chart_in_pdf(elements, req.chart_image)
    
    # Table Data
    data = [['Invoice No', 'Date', 'Supplier', 'Medicines', 'Total Qty', 'Grand Total', 'Status']]
    for t in report_data.transactions:
        data.append([
            t.InvoiceNo, 
            t.PurchaseDate.strftime("%Y-%m-%d %H:%M"),
            t.SupplierName[:15],
            str(t.MedicinesPurchased),
            str(t.TotalQty),
            str(round(t.GrandTotal, 2)),
            t.Status
        ])
        
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#FFFFFF')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 12),
        ('BOTTOMPADDING', (0,0), (-1,0), 14),
        ('TOPPADDING', (0,0), (-1,0), 14),
        
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#FFFFFF'), colors.HexColor('#F1F5F9')]),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-1), 10),
        ('BOTTOMPADDING', (0,1), (-1,-1), 10),
        ('TOPPADDING', (0,1), (-1,-1), 10),
        
        ('LINEBELOW', (0,0), (-1,0), 2, colors.HexColor('#3B82F6')), 
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('BOX', (0,0), (-1,-1), 1.5, colors.HexColor('#94A3B8')),
    ]))
    elements.append(t)
    
    doc.build(elements)
    response = Response(content=output.getvalue(), media_type="application/pdf")
    response.headers["Content-Disposition"] = f"attachment; filename=purchase_report_{sd}_to_{ed}.pdf"
    return response

@router.post("/inventory/export/pdf")
def export_inventory_report_pdf(req: dict = Body(...), db: Session = Depends(get_db)):
    req = PDFExportRequest(**req)
    sd, ed = get_reports_date_range(req.timeframe, req.start_date, req.end_date)
    report_data = fetch_inventory_report_data(db, sd, ed)
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    elements.extend(build_premium_header("Inventory Report", f"As of: {ed}"))
    
    kpi_data = [
        ("Total Items", str(report_data.summary.TotalItemsInStock), "#6366F1"),
        ("Cost Value", f"Rs. {report_data.summary.TotalCostValue}", "#3B82F6"),
        ("Low Stock", str(report_data.summary.LowStockCount), "#F59E0B"),
        ("Out of Stock", str(report_data.summary.OutOfStockCount), "#EF4444"),
    ]
    elements.extend(build_kpi_table(kpi_data))
    
    embed_chart_in_pdf(elements, req.chart_image)
    
    data = [['Medicine Name', 'Category', 'Batch', 'Stock Qty', 'Cost Price', 'Selling Price', 'Status']]
    for t in report_data.stock_items[:200]: # limit to 200 items to avoid giant PDF
        data.append([
            t.MedicineName[:15],
            t.Category[:10],
            t.BatchCode,
            str(t.Quantity),
            str(round(t.CostPrice, 2)),
            str(round(t.SellingPrice, 2)),
            t.Status
        ])
        
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#FFFFFF')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 12),
        ('BOTTOMPADDING', (0,0), (-1,0), 14),
        ('TOPPADDING', (0,0), (-1,0), 14),
        
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#FFFFFF'), colors.HexColor('#F1F5F9')]),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-1), 10),
        ('BOTTOMPADDING', (0,1), (-1,-1), 10),
        ('TOPPADDING', (0,1), (-1,-1), 10),
        
        ('LINEBELOW', (0,0), (-1,0), 2, colors.HexColor('#3B82F6')), 
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('BOX', (0,0), (-1,-1), 1.5, colors.HexColor('#94A3B8')),
    ]))
    elements.append(t)
    if len(report_data.stock_items) > 200:
        elements.append(Paragraph(f"(Showing first 200 records out of {len(report_data.stock_items)}. Export to Excel for full list.)", styles['Normal']))
    
    doc.build(elements)
    response = Response(content=output.getvalue(), media_type="application/pdf")
    response.headers["Content-Disposition"] = f"attachment; filename=inventory_report.pdf"
    return response

@router.post("/medicine/export/pdf")
def export_medicine_report_pdf(req: dict = Body(...), db: Session = Depends(get_db)):
    req = PDFExportRequest(**req)
    sd, ed = get_reports_date_range(req.timeframe, req.start_date, req.end_date)
    # Using page_size=0 to ensure we have all data without limit
    report_data = fetch_medicine_report_data(db, sd, ed, req.report_type, None, None, 1, 0)
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    title_map = {
        'expiry': 'Medicine Expiry Alerts',
        'low_stock': 'Medicine Low Stock Report',
        'moving': 'Medicine Performance Report'
    }
    
    elements.extend(build_premium_header(title_map.get(req.report_type, "Medicine Report"), f"Period: {sd} to {ed}" if req.report_type == 'moving' else f"Snapshot As Of: {ed}"))
    
    kpi_data = [
        ("Total Expired Batches", str(report_data.summary.TotalExpiredBatches), "#EF4444"),
        ("Expiring Soon (90d)", str(report_data.summary.ExpiringSoonBatches), "#F59E0B"),
        ("Low Stock", str(report_data.summary.LowStockMedicines), "#6366F1"),
        ("Fast Moving", str(report_data.summary.FastMovingCount), "#10B981"),
    ]
    elements.extend(build_kpi_table(kpi_data))
    
    embed_chart_in_pdf(elements, req.chart_image)
    
    if req.report_type == 'expiry':
        data = [['Medicine', 'Batch', 'Supplier', 'Qty', 'Expiry Date', 'Days', 'Status']]
        for t in report_data.expiry_items[:200]:
            data.append([
                t.MedicineName[:15],
                t.BatchCode,
                t.SupplierName[:15] if getattr(t, 'SupplierName', None) else '-',
                str(t.Quantity),
                t.ExpiryDate.strftime("%Y-%m-%d") if getattr(t, 'ExpiryDate', None) else 'N/A',
                str(t.DaysToExpiry),
                t.Status
            ])
    elif req.report_type == 'low_stock':
        data = [['Medicine', 'Category', 'Supplier', 'Stock', 'Reorder Level', 'Deficit', 'Suggested']]
        for t in report_data.low_stock_items[:200]:
            data.append([
                t.MedicineName[:15],
                t.Category[:10],
                t.SupplierName[:15] if getattr(t, 'SupplierName', None) else '-',
                str(getattr(t, 'CurrentStock', 0)),
                str(getattr(t, 'ReorderLevel', 0)),
                str(getattr(t, 'Deficit', 0)),
                str(getattr(t, 'SuggestedReorderQty', 0))
            ])
    else:
        data = [['Medicine', 'Category', 'Supplier', 'Qty Sold', 'Velocity/Day', 'Revenue', 'Classification']]
        for t in report_data.movement_items[:200]:
            data.append([
                t.MedicineName[:15],
                t.Category[:10],
                t.SupplierName[:15] if getattr(t, 'SupplierName', None) else '-',
                str(getattr(t, 'SoldQuantity', 0)),
                str(round(getattr(t, 'SalesVelocity', 0.0), 2)),
                str(round(getattr(t, 'Revenue', 0.0), 2)),
                getattr(t, 'Classification', 'Unknown')
            ])
            
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#FFFFFF')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 12),
        ('BOTTOMPADDING', (0,0), (-1,0), 14),
        ('TOPPADDING', (0,0), (-1,0), 14),
        
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#FFFFFF'), colors.HexColor('#F1F5F9')]),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-1), 10),
        ('BOTTOMPADDING', (0,1), (-1,-1), 10),
        ('TOPPADDING', (0,1), (-1,-1), 10),
        
        ('LINEBELOW', (0,0), (-1,0), 2, colors.HexColor('#3B82F6')), 
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('BOX', (0,0), (-1,-1), 1.5, colors.HexColor('#94A3B8')),
    ]))
    elements.append(t)
    
    if (req.report_type == 'expiry' and len(report_data.expiry_items) > 200) or \
       (req.report_type == 'low_stock' and len(report_data.low_stock_items) > 200) or \
       (req.report_type == 'moving' and len(report_data.movement_items) > 200):
        elements.append(Paragraph("(Showing first 200 records. Export to CSV for full list.)", styles['Normal']))
    
    doc.build(elements)
    response = Response(content=output.getvalue(), media_type="application/pdf")
    response.headers["Content-Disposition"] = f"attachment; filename=medicine_{req.report_type}_report.pdf"
    return response

@router.post("/financial/export/pdf")
def export_financial_report_pdf(req: dict = Body(...), db: Session = Depends(get_db)):
    req = PDFExportRequest(**req)
    sd, ed = get_reports_date_range(req.timeframe, req.start_date, req.end_date)
    report_data = fetch_financial_report_data(db, sd, ed)
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    elements.extend(build_premium_header("Profit & Loss Statement", f"Period: {sd} to {ed}"))
    
    kpi_data = [
        ("Net Revenue", f"Rs. {report_data.summary.TotalRevenue}", "#3B82F6"),
        ("COGS", f"Rs. {report_data.summary.TotalCOGS}", "#F59E0B"),
        ("Gross Profit", f"Rs. {report_data.summary.GrossProfit}", "#6366F1"),
        ("Net Profit", f"Rs. {report_data.summary.NetProfit}", "#10B981" if report_data.summary.NetProfit >= 0 else "#EF4444"),
    ]
    elements.extend(build_kpi_table(kpi_data))
    
    embed_chart_in_pdf(elements, req.chart_image)
    
    data = [['Description', 'Amount (Rs)', 'Section']]
    
    data.append(["Gross Sales Revenue", f"+ {round(report_data.summary.GrossSales, 2)}", "1. Revenue (Income)"])
    data.append(["Less: Sales Returns & Refunds", f"- {round(report_data.summary.SalesReturns, 2)}", "1. Revenue (Income)"])
    data.append(["Less: Discounts Given", f"- {round(report_data.summary.DiscountsApplied, 2)}", "1. Revenue (Income)"])
    data.append(["Subtotal: Net Revenue", f"{round(report_data.summary.TotalRevenue, 2)}", "1. Revenue (Income)"])

    data.append(["Direct Cost of Sold Medicines", f"- {round(report_data.summary.TotalCOGS, 2)}", "2. Cost of Goods Sold (COGS)"])
    data.append(["Subtotal: Gross Profit", f"{round(report_data.summary.GrossProfit, 2)}", "2. Cost of Goods Sold (COGS)"])

    data.append(["Inventory Expiry & Write-Offs", f"- {round(report_data.summary.InventoryLoss, 2)}", "3. Expenses & Losses"])
    data.append(["Operating Expenses", f"- {round(report_data.summary.TotalExpenses, 2)}", "3. Expenses & Losses"])

    data.append(["NET PROFIT / LOSS", f"{round(report_data.summary.NetProfit, 2)}", "4. Final Summary"])
        
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#FFFFFF')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ALIGN', (0,1), (0,-1), 'LEFT'),  # Left align descriptions
        ('ALIGN', (1,1), (1,-1), 'RIGHT'), # Right align amounts
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 12),
        ('BOTTOMPADDING', (0,0), (-1,0), 14),
        ('TOPPADDING', (0,0), (-1,0), 14),
        
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#FFFFFF'), colors.HexColor('#F1F5F9')]),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-1), 10),
        ('BOTTOMPADDING', (0,1), (-1,-1), 10),
        ('TOPPADDING', (0,1), (-1,-1), 10),
        
        # Bold subtotals and Net Profit
        ('FONTNAME', (0, 4), (-1, 4), 'Helvetica-Bold'),
        ('FONTNAME', (0, 6), (-1, 6), 'Helvetica-Bold'),
        ('FONTNAME', (0, 9), (-1, 9), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 9), (-1, 9), 12),
        
        ('LINEBELOW', (0,0), (-1,0), 2, colors.HexColor('#3B82F6')), 
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('BOX', (0,0), (-1,-1), 1.5, colors.HexColor('#94A3B8')),
    ]))
    elements.append(t)
    
    doc.build(elements)
    response = Response(content=output.getvalue(), media_type="application/pdf")
    response.headers["Content-Disposition"] = f"attachment; filename=financial_report.pdf"
    return response



@router.get("/financial/export/csv")
def export_financial_report_csv(
    timeframe: str = 'this_month',
    start_date: str = None,
    end_date: str = None,
    db: Session = Depends(get_db)
):
    sd, ed = get_reports_date_range(timeframe, start_date, end_date)
    report_data = fetch_financial_report_data(db, sd, ed)
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Category', 'Amount', 'Type'])
    
    for item in report_data.income_breakdown:
        writer.writerow([item.Category, round(item.Amount, 2), 'Income'])
    for item in report_data.expense_breakdown:
        writer.writerow([item.Category, round(item.Amount, 2), 'Expense'])
        
    writer.writerow([])
    writer.writerow(['Summary', 'Amount'])
    writer.writerow(['Total Revenue', round(report_data.summary.TotalRevenue, 2)])
    writer.writerow(['Total COGS', round(report_data.summary.TotalCOGS, 2)])
    writer.writerow(['Gross Profit', round(report_data.summary.GrossProfit, 2)])
    writer.writerow(['Operating Expenses', round(report_data.summary.OperatingExpenses, 2)])
    writer.writerow(['Net Profit', round(report_data.summary.NetProfit, 2)])
    writer.writerow(['Profit Margin', f'{round(report_data.summary.ProfitMargin, 2)}%'])
    
    response = Response(content=output.getvalue(), media_type="text/csv")
    response.headers["Content-Disposition"] = f"attachment; filename=financial_report.csv"
    return response
